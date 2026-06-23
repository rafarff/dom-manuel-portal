#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador do PORTAL DO CLIENTE — Condomínio Dom Manuel
=====================================================
Parte voltada à apresentação para clientes (os sócios abrem e vendem).
Gera um único `index.html` autossuficiente com:
  - portão de senha única (cortina);
  - aba BOOK  -> iframe para book.html + botão "Apresentar em tela cheia";
  - aba ESPELHO -> grade 15×3 com status + R$/m² por unidade;
  - aba TABELA VIGENTE -> layout clássico (3 blocos Final/Tipo) com as
    colunas de parcelas, igual ao Excel/PDF por tabela.

A parte INTERNA (Status Geral, plano comercial, demais tabelas) será um
segundo arquivo (gestao.html), feito depois.

FLUXO DE ATUALIZAÇÃO:
  1. Atualize o painel (_Controle_Comercial/...xlsx)  →  fonte de status.
  2. Rode:  python3 gerar_portal.py
  3. Publique a pasta no GitHub Pages (commit + push).

CONFIG abaixo: senha e tabela vigente.
"""

import os, html
from openpyxl import load_workbook

# ───────────────────────── CONFIG ─────────────────────────
HERE = os.path.dirname(os.path.abspath(__file__))
COMERCIAL = os.path.abspath(os.path.join(HERE, ".."))
PAINEL = os.path.join(COMERCIAL, "_Controle_Comercial", "01_Controle_Comercial_Dom_Manuel.xlsx")
REGUA  = os.path.join(COMERCIAL, "01_Tabela_Comercial", "01_Tabela_Comercial_Dom_Manuel_v3.xlsx")
OUT    = os.path.join(HERE, "index.html")

CONFIG = {
    "senha": "1482",            # ← senha única do portal (troque aqui)
    "vigente": "T1",            # ← tabela vigente exibida ao cliente
    "subtitulo": "Ponta D'Areia · São Luís/MA",
    "tema": "escuro",           # ← "escuro" | "contraste" | "claro_sutil" | "claro"
    "mostrar_vendido": False,   # ← False nesta fase: permuta/vendida viram "Reservado" (omite RI)
    "regua_versao": "v3.3",     # ← rótulo de versão no cabeçalho do PDF exportado
}

# ───────────────────────── TEMAS ─────────────────────────
# Cada tema define a base (preto/grafite/dourado etc.) e o tratamento de status
# (av = disponível · rs = reservado · sd = vendido). O foco é evidência de status.
TEMAS = {
    # 1) Escuro sutil — original (bordas coloridas, fundo discreto)
    "escuro": {
        "bg": "#16140f", "panel": "#1f1c16", "panel2": "#26221b", "gold": "#c2a15a",
        "golds": "#d9c08a", "line": "#3a342a", "txt": "#ece6da", "muted": "#9a9285", "green": "#3D4E45",
        "av-bg": "#1b2a20", "av-bd": "#33503f", "av-tx": "#7fc69a",
        "rs-bg": "#2a2213", "rs-bd": "#6d5a2c", "rs-tx": "#e0b552",
        "sd-bg": "#241818", "sd-bd": "#5a3f3f", "sd-tx": "#cc9a9a",
    },
    # 2) Contraste — escuro com PREENCHIMENTO sólido (status saltam)
    "contraste": {
        "bg": "#14130f", "panel": "#1e1b15", "panel2": "#26221b", "gold": "#cda85f",
        "golds": "#e6cf94", "line": "#3a342a", "txt": "#f1ecdf", "muted": "#9a9285", "green": "#3D4E45",
        "av-bg": "#1f7a4d", "av-bd": "#2fa468", "av-tx": "#f2fff7",
        "rs-bg": "#b5642f", "rs-bd": "#d98640", "rs-tx": "#fff4ea",
        "sd-bg": "#7a3b3b", "sd-bd": "#9b5252", "sd-tx": "#ffeaea",
    },
    # 3b) Claro sutil — fundo creme, status em tons translúcidos (arejado)
    "claro_sutil": {
        "bg": "#f6f2ea", "panel": "#fffdf9", "panel2": "#f0ebe0", "gold": "#9c7b2e",
        "golds": "#7d6021", "line": "#e4dccb", "txt": "#2f2a1d", "muted": "#8a8270", "green": "#3D4E45",
        "av-bg": "rgba(31,122,77,0.12)", "av-bd": "rgba(31,122,77,0.55)", "av-tx": "#1c6a40",
        "rs-bg": "rgba(194,118,47,0.15)", "rs-bd": "rgba(194,118,47,0.60)", "rs-tx": "#955310",
        "sd-bg": "rgba(155,82,82,0.14)", "sd-bd": "rgba(155,82,82,0.55)", "sd-tx": "#8a3b3b",
    },
    # 3) Claro — fundo creme, status em chips fortes
    "claro": {
        "bg": "#f3eee3", "panel": "#ffffff", "panel2": "#efe9dc", "gold": "#9c7b2e",
        "golds": "#7d6021", "line": "#e2d9c6", "txt": "#2c2719", "muted": "#857c68", "green": "#3D4E45",
        "av-bg": "#1f7a4d", "av-bd": "#1f7a4d", "av-tx": "#ffffff",
        "rs-bg": "#c2762f", "rs-bd": "#c2762f", "rs-tx": "#ffffff",
        "sd-bg": "#9b5252", "sd-bd": "#9b5252", "sd-tx": "#ffffff",
    },
}

# ───────────────────────── FORMATAÇÃO ─────────────────────────
def fmt2(v):
    try: v = float(v)
    except (TypeError, ValueError): return "—"
    return "R$ " + f"{v:,.2f}".replace(",", "#").replace(".", ",").replace("#", ".")

def fmt0(v):
    try: v = float(v)
    except (TypeError, ValueError): return "—"
    return "R$ " + f"{v:,.0f}".replace(",", ".")

def rm2_fmt(v):
    try: v = float(v)
    except (TypeError, ValueError): return ""
    return f"{v:,.0f}".replace(",", ".") + "/m²"

def esc(s): return html.escape(str(s)) if s is not None else ""

AREA = {"01": 134.68, "02": 116.38, "03": 134.68}  # Tipo 01 (finais 01/03) · Tipo 02 (final 02)

def media_rm2_fmt(tab):
    """Preço médio R$/m² ponderado pela área (sobre o preço total) da tabela vigente."""
    tv = ta = 0.0
    for b in tab["blocks"]:
        for row in b["rows"]:
            tv += float(row["valor"]); ta += AREA.get(row["apto"][-2:], 0)
    m = tv / ta if ta else 0
    return "R$ " + f"{m:,.0f}".replace(",", ".") + "/m²"

# ───────────────────────── LEITURA ─────────────────────────
def ler_painel():
    wb = load_workbook(PAINEL, data_only=True)
    ws = wb["Comercial"]
    rows = list(ws.iter_rows(values_only=True))
    status_by_apto = {}
    atual = ""
    for cell in (rows[2][0] or "").split("·"):
        if "Atualizado" in cell:
            atual = cell.replace("Atualizado em", "").strip()
    for r in rows[5:]:
        if not r or r[1] in (None, ""):
            continue
        status_by_apto[str(r[1]).strip()] = (r[5] or "").strip()
    return status_by_apto, atual

def mascara(status):
    """status do painel -> rótulo externo. Nesta fase (mostrar_vendido=False)
    permuta/vendida também aparecem como 'Reservado' — omite o RI ao cliente."""
    s = (status or "").lower()
    if "permuta" in s or "vendid" in s:
        return "Vendido" if CONFIG["mostrar_vendido"] else "Reservado"
    if "contrato" in s or "reserv" in s:
        return "Reservado"
    return "Disponível"

def ler_tabela(key):
    """Lê a aba TX da régua -> blocos (Final/Tipo) com colunas de parcelas + observações."""
    wb = load_workbook(REGUA, data_only=True)
    ws = wb[key]
    blocks, obs = [], []
    cur, in_obs = None, False
    titulo = ""
    for r in ws.iter_rows(values_only=True):
        c0 = r[0]
        s0 = str(c0).strip() if c0 is not None else ""
        if s0.upper().startswith("TABELA COMERCIAL"):
            titulo = s0; continue
        if s0.upper().startswith("FINAL"):
            cur = {"label": s0, "rows": []}; blocks.append(cur); in_obs = False; continue
        if s0.upper() in ("OBSERVAÇÕES", "OBSERVACOES"):
            in_obs = True; cur = None; continue
        if in_obs:
            if s0 and not s0.upper().startswith("DOM INCORP"):
                obs.append(s0)
            continue
        if cur is not None and isinstance(c0, (int, float)):
            cur["rows"].append({
                "andar": int(c0), "apto": str(r[1]).strip(),
                "valor": r[3], "sinal": r[4], "parcela": r[5],
                "reforco": r[6], "fin": r[7], "rm2": r[8],
            })
    # ordenar blocos: Final 01 -> 02 -> 03
    import re
    def _ord(b):
        m = re.search(r'FINAL\s*0?(\d)', b["label"].upper())
        return int(m.group(1)) if m else 9
    blocks.sort(key=_ord)
    # índice apto -> dados (p/ espelho)
    idx = {}
    for b in blocks:
        for row in b["rows"]:
            idx[row["apto"]] = row
    return {"titulo": titulo, "blocks": blocks, "obs": obs, "idx": idx}

# ───────────────────────── COMPONENTES HTML ─────────────────────────
def espelho_html(status_by_apto, idx):
    cls_map = {"Disponível": "e-ok", "Reservado": "e-res", "Vendido": "e-sold"}
    rows_html = []
    for andar in range(15, 0, -1):
        cells = []
        for final in ["01", "02", "03"]:
            apto = f"{andar}{final}"
            rotulo = mascara(status_by_apto.get(apto, "Disponível"))
            cls = cls_map.get(rotulo, "e-ok")
            r = idx.get(apto, {})
            preco_html = ""
            if r.get("rm2"):
                preco_html = (f'<span class="erm2 pr-rm2">R$ {esc(rm2_fmt(r["rm2"]))}</span>'
                              f'<span class="erm2 eval pr-ticket">{esc(fmt0(r["valor"]))}</span>')
            cells.append(
                f'<td class="ecell {cls}"><span class="eapto">{apto}</span>'
                f'<span class="estat">{rotulo}</span>{preco_html}</td>'
            )
        rows_html.append(f'<tr><th class="eand">{andar}º</th>{"".join(cells)}</tr>')
    head = ('<tr><th></th><th>Final 01<small>lado lagoa</small></th>'
            '<th>Final 02<small>miolo</small></th>'
            '<th>Final 03<small>lado mar</small></th></tr>')
    return f'<table class="espelho">{head}{"".join(rows_html)}</table>'

def folha_espelho(status_by_apto, idx, vigente, atual, regua_versao, media):
    """Bloco DEDICADO de exportação (A4 retrato, células preenchidas) — fora da tela."""
    cls = {"Disponível": "f-ok", "Reservado": "f-res", "Vendido": "f-sold"}
    rows = []
    for andar in range(15, 0, -1):
        tds = []
        for final in ["01", "02", "03"]:
            apto = f"{andar}{final}"
            rot = mascara(status_by_apto.get(apto, "Disponível"))
            c = cls.get(rot, "f-ok")
            r = idx.get(apto, {})
            pm = ""
            if r.get("rm2"):
                pm = (f'<span class="fm pr-rm2">R$ {esc(rm2_fmt(r["rm2"]))}</span>'
                      f'<span class="fm fmv pr-ticket">{esc(fmt0(r["valor"]))}</span>')
            tds.append(f'<td class="fcell {c}"><span class="fa">{apto}</span>'
                       f'<span class="fs">{rot}</span>{pm}</td>')
        rows.append(f'<tr><td class="fand">{andar}º</td>{"".join(tds)}</tr>')
    head = ('<tr><th></th><th>Final 01<small>lado lagoa</small></th>'
            '<th>Final 02<small>miolo</small></th><th>Final 03<small>lado mar</small></th></tr>')
    return (
        '<div class="folha-head"><div><div class="fh-brand">DOM <b>MANUEL</b></div>'
        "<div class=\"fh-sub\">Condomínio Dom Manuel · Ponta D'Areia — São Luís/MA</div></div>"
        f'<div class="fh-right"><div class="fh-tt">Espelho de Vendas · {esc(vigente)}</div>'
        f'<div class="fh-media">Preço médio {esc(media)}</div>'
        f'<div class="fh-dt">Régua {esc(regua_versao)} · atualizado {esc(atual or "—")}</div></div></div>'
        f'<table class="fesp">{head}{"".join(rows)}</table>'
        '<div class="folha-legend"><span><i style="background:rgba(31,122,77,0.15);border:1px solid rgba(31,122,77,0.45)"></i>Disponível</span>'
        '<span><i style="background:rgba(194,118,47,0.17);border:1px solid rgba(194,118,47,0.5)"></i>Reservado</span></div>'
        '<div class="folha-foot">Espelho para fins de reserva, sujeito a alteração sem aviso · '
        'valores sujeitos a análise de crédito · plantas e imagens ilustrativas · DOM Incorporação</div>'
    )

def folha_tabela(tab, status_by_apto, vigente, atual, regua_versao, media):
    """Bloco DEDICADO de exportação da tabela (A4 paisagem, papel branco)."""
    cls = {"Disponível": "p-ok", "Reservado": "p-res", "Vendido": "p-res"}
    blocks = ""
    for b in tab["blocks"]:
        rows = ""
        for row in b["rows"]:
            rot = mascara(status_by_apto.get(row["apto"], "Disponível"))
            pc = cls.get(rot, "p-ok")
            rows += (f'<tr><td class="l">{row["andar"]}º</td><td class="l b">{esc(row["apto"])}</td>'
                     f'<td class="l"><span class="pill {pc}">{rot}</span></td>'
                     f'<td>R$ {esc(rm2_fmt(row["rm2"]))}</td><td>{fmt2(row["sinal"])}</td><td>{fmt2(row["parcela"])}</td>'
                     f'<td>{fmt2(row["reforco"])}</td><td>{fmt2(row["fin"])}</td>'
                     f'<td>{fmt2(row["valor"])}</td></tr>')
        head = ('<tr><th class="l">And.</th><th class="l">Apto</th><th class="l">Status</th>'
                '<th>R$/m²</th><th>Sinal (ato)</th><th>Parc. mensal (40×)</th>'
                '<th>Reforço (3×)</th><th>Financ. (habite-se)</th><th>Valor total</th></tr>')
        blocks += f'<div class="ftbl-blk">{esc(b["label"])}</div><table class="ftbl">{head}{rows}</table>'
    obs = ""
    if tab["obs"]:
        lis = "".join(f"<div>• {esc(o)}</div>" for o in tab["obs"])
        obs = f'<div class="ftbl-obs"><b>Observações</b>{lis}</div>'
    return (
        '<div class="folha-head"><div><div class="fh-brand">DOM <b>MANUEL</b></div>'
        "<div class=\"fh-sub\">Condomínio Dom Manuel · Ponta D'Areia — São Luís/MA</div></div>"
        f'<div class="fh-right"><div class="fh-tt">Tabela Comercial · {esc(vigente)}</div>'
        f'<div class="fh-media">Preço médio {esc(media)}</div>'
        f'<div class="fh-dt">Régua {esc(regua_versao)} · atualizado {esc(atual or "—")}</div></div></div>'
        + blocks + obs
    )

def tabela_classica(tab, status_by_apto):
    cls_map = {"Disponível": "st-ok", "Reservado": "st-res", "Vendido": "st-sold"}
    blocks_html = []
    for b in tab["blocks"]:
        rows_html = []
        for row in b["rows"]:
            rotulo = mascara(status_by_apto.get(row["apto"], "Disponível"))
            cls = cls_map.get(rotulo, "st-ok")
            rows_html.append(
                f'<tr class="{cls}"><td class="l">{row["andar"]}º</td>'
                f'<td class="l b">{esc(row["apto"])}</td>'
                f'<td class="l"><span class="pill">{rotulo}</span></td>'
                f'<td class="r">R$ {esc(rm2_fmt(row["rm2"]))}</td>'
                f'<td class="r">{fmt2(row["sinal"])}</td>'
                f'<td class="r">{fmt2(row["parcela"])}</td>'
                f'<td class="r">{fmt2(row["reforco"])}</td>'
                f'<td class="r">{fmt2(row["fin"])}</td>'
                f'<td class="r">{fmt2(row["valor"])}</td></tr>'
            )
        head = ('<tr><th>And.</th><th>Apto</th><th>Status</th><th>R$/m²</th>'
                '<th>Sinal (ato)</th><th>Parc. mensal (40×)</th><th>Reforço (3×)</th>'
                '<th>Financ. (habite-se)</th><th>Valor total</th></tr>')
        blocks_html.append(
            f'<div class="blocohdr">{esc(b["label"])}</div>'
            f'<table class="ptable">{head}{"".join(rows_html)}</table>'
        )
    obs_html = ""
    if tab["obs"]:
        lis = "".join(f"<li>{esc(o)}</li>" for o in tab["obs"])
        obs_html = f'<div class="obs"><b>Observações</b><ul>{lis}</ul></div>'
    return "".join(blocks_html) + obs_html

# ───────────────────────── RENDER ─────────────────────────
def render(status_by_apto, atual, tab, vigente, C):
    esp = espelho_html(status_by_apto, tab["idx"])
    tabela = tabela_classica(tab, status_by_apto)
    media = media_rm2_fmt(tab)
    folha_esp = folha_espelho(status_by_apto, tab["idx"], vigente, atual, CONFIG["regua_versao"], media)
    folha_tab = folha_tabela(tab, status_by_apto, vigente, atual, CONFIG["regua_versao"], media)
    root_vars = ";".join(f"--{k}:{v}" for k, v in C.items())
    mostrar_sd = CONFIG["mostrar_vendido"]
    legenda = ('<span><i style="background:var(--av-bg);border:1px solid var(--av-bd)"></i>Disponível</span>'
               '<span><i style="background:var(--rs-bg);border:1px solid var(--rs-bd)"></i>Reservado</span>')
    if mostrar_sd:
        legenda += '<span><i style="background:var(--sd-bg);border:1px solid var(--sd-bd)"></i>Vendido</span>'
    return f"""<!DOCTYPE html>
<html lang="pt-BR"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Dom Manuel · Apresentação</title>
<style>
:root{{{root_vars}}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:var(--bg);color:var(--txt);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;line-height:1.5}}
a{{color:var(--golds);text-decoration:none}}
#gate{{position:fixed;inset:0;z-index:9999;background:var(--bg);display:flex;align-items:center;justify-content:center;flex-direction:column;gap:16px}}
#gate .brand{{font-size:32px;letter-spacing:.2em;color:var(--gold);font-weight:300}}
#gate .sub{{color:var(--muted);font-size:12px;letter-spacing:.18em}}
#gate form{{display:flex;gap:8px;margin-top:12px}}
#gate input{{background:var(--panel);border:1px solid var(--line);color:var(--txt);padding:12px 16px;border-radius:8px;font-size:15px;min-width:220px}}
#gate button{{background:var(--gold);color:#1a1710;border:0;padding:12px 22px;border-radius:8px;font-weight:600;cursor:pointer}}
#gate .err{{color:#d98c8c;font-size:13px;height:16px}}
#app{{display:none;max-width:1080px;margin:0 auto;padding:0 18px 80px}}
header.top{{display:flex;justify-content:space-between;align-items:flex-end;padding:26px 0 16px;border-bottom:1px solid var(--line)}}
header.top .brand{{font-size:24px;letter-spacing:.14em;color:var(--gold);font-weight:300}}
header.top .brand b{{font-weight:600}}
header.top .meta{{text-align:right;color:var(--muted);font-size:12px;line-height:1.7}}
nav.tabs{{display:flex;gap:4px;flex-wrap:wrap;position:sticky;top:0;background:var(--bg);padding:14px 0;z-index:50;border-bottom:1px solid var(--line)}}
nav.tabs button{{background:transparent;border:1px solid transparent;color:var(--muted);padding:9px 16px;border-radius:8px;font-size:14px;cursor:pointer;letter-spacing:.03em}}
nav.tabs button.on{{color:var(--gold);border-color:var(--line);background:var(--panel)}}
nav.tabs button.soon{{opacity:.45;cursor:default}}
section.view{{display:none;padding-top:22px}}
section.view.on{{display:block}}
h2.vt{{font-size:13px;letter-spacing:.16em;text-transform:uppercase;color:var(--gold);font-weight:600;margin-bottom:4px}}
p.vd{{color:var(--muted);font-size:13px;margin-bottom:16px}}
.btnrow{{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px}}
.act{{background:var(--gold);color:#1a1710;border:0;padding:11px 20px;border-radius:8px;font-weight:600;font-size:14px;cursor:pointer}}
.act.ghost{{background:transparent;color:var(--golds);border:1px solid var(--line)}}
.bookframe{{width:100%;height:74vh;border:1px solid var(--line);border-radius:12px;background:var(--green)}}
/* espelho */
table.espelho{{width:100%;border-collapse:separate;border-spacing:6px}}
table.espelho th{{color:var(--muted);font-weight:500;font-size:12px;text-align:center;padding:4px}}
table.espelho th small{{display:block;color:var(--muted);opacity:.6;font-size:10px}}
th.eand{{color:var(--golds);font-size:13px}}
.ecell{{border-radius:8px;text-align:center;padding:9px 4px;border:1px solid var(--line);background:var(--panel)}}
.ecell .eapto{{display:block;font-weight:600;font-size:14px}}
.ecell .estat{{display:block;font-size:10.5px;letter-spacing:.03em;margin-top:2px}}
.ecell .erm2{{display:block;font-size:11px;margin-top:3px;font-weight:600}}
.e-ok{{background:var(--av-bg);border-color:var(--av-bd)}}.e-ok .eapto,.e-ok .estat,.e-ok .erm2{{color:var(--av-tx)}}
.e-res{{background:var(--rs-bg);border-color:var(--rs-bd)}}.e-res .eapto,.e-res .estat,.e-res .erm2{{color:var(--rs-tx)}}
.e-sold{{background:var(--sd-bg);border-color:var(--sd-bd)}}.e-sold .eapto,.e-sold .estat,.e-sold .erm2{{color:var(--sd-tx)}}
.ecell .erm2{{opacity:.9}}
.legend{{display:flex;gap:18px;flex-wrap:wrap;margin:16px 0 4px;font-size:12px;color:var(--muted)}}
.legend i{{display:inline-block;width:11px;height:11px;border-radius:3px;margin-right:6px;vertical-align:middle}}
/* tabela clássica */
.blocohdr{{background:linear-gradient(90deg,var(--green),transparent);border-left:3px solid var(--gold);color:var(--golds);font-size:12.5px;letter-spacing:.04em;padding:9px 12px;margin:22px 0 0;border-radius:4px}}
table.ptable{{width:100%;border-collapse:collapse;font-size:12.5px;margin-bottom:6px}}
table.ptable th{{background:var(--panel2);color:var(--muted);font-weight:500;font-size:10.5px;letter-spacing:.04em;text-transform:uppercase;padding:8px 7px;text-align:right;border-bottom:1px solid var(--line)}}
table.ptable th:first-child,table.ptable th:nth-child(2),table.ptable th:nth-child(3){{text-align:left}}
table.ptable td{{padding:7px 7px;border-bottom:1px solid var(--line);font-variant-numeric:tabular-nums}}
table.ptable td.l{{text-align:left}}table.ptable td.r{{text-align:right}}td.b{{font-weight:600}}
.pill{{font-size:10.5px;padding:2px 9px;border-radius:20px;border:1px solid var(--line);font-weight:600}}
tr.st-ok .pill{{color:var(--av-tx);background:var(--av-bg);border-color:var(--av-bd)}}
tr.st-res .pill{{color:var(--rs-tx);background:var(--rs-bg);border-color:var(--rs-bd)}}
tr.st-sold .pill{{color:var(--sd-tx);background:var(--sd-bg);border-color:var(--sd-bd)}}
.obs{{margin-top:20px;background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:14px 18px;font-size:11.5px;color:var(--muted)}}
.obs b{{color:var(--golds);display:block;margin-bottom:6px;font-size:11px;letter-spacing:.08em;text-transform:uppercase}}
.obs ul{{padding-left:18px}}.obs li{{margin:3px 0}}
.printbtn{{background:transparent;border:1px solid var(--line);color:var(--muted);padding:7px 14px;border-radius:8px;font-size:12px;cursor:pointer;float:right}}
.toggleprice{{background:var(--panel);border:1px solid var(--line);color:var(--gold);padding:8px 16px;border-radius:8px;font-size:13px;cursor:pointer;margin-bottom:14px}}
.toggleprice.on{{background:var(--gold);color:#1a1710;border-color:var(--gold)}}
.ecell .pr-rm2,.ecell .pr-ticket,.fcell .pr-rm2,.fcell .pr-ticket{{display:none}}
body.show-rm2 .ecell .pr-rm2,body.show-rm2 .fcell .pr-rm2{{display:block}}
body.show-ticket .ecell .pr-ticket,body.show-ticket .fcell .pr-ticket{{display:block}}
.ecell .eval{{font-size:10px;opacity:.8}} .fcell .fmv{{font-size:7px;margin-top:0}}
.soonbox{{background:var(--panel);border:1px dashed var(--line);border-radius:12px;padding:40px;text-align:center;color:var(--muted);font-size:14px}}
footer{{color:var(--line);font-size:11px;text-align:center;margin-top:40px}}
.printhdr{{display:none;justify-content:space-between;align-items:flex-end;background:#16140f;padding:9px 14px;margin-bottom:10px;border-radius:6px}}
.printhdr .ph-brand{{font-size:15pt;font-weight:300;letter-spacing:.12em;color:#c2a15a}}.printhdr .ph-brand b{{font-weight:600}}
.printhdr .ph-sub{{font-size:8pt;color:#d9c08a}}
.printhdr .ph-right{{text-align:right}}.printhdr .ph-right span{{display:block;font-size:11pt;color:#c2a15a;font-weight:600}}.printhdr .ph-right small{{font-size:7pt;color:#9a9285}}
.printftr{{display:none;margin-top:12px;font-size:7pt;color:#6a6253;text-align:center;border-top:.5px solid #cfc6b3;padding-top:6px}}
/* ---- FOLHA de exportação dedicada (bloco branco, fora da tela) ---- */
#folhas{{display:none}}
.folha{{background:#fff;color:#2c2719;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif}}
#folha-espelho{{max-width:620px;margin:0 auto}}
.folha-head{{display:flex;justify-content:space-between;align-items:flex-end;background:transparent;border-bottom:2.5px solid #b9974a;padding:0 2px 7px}}
.folha-head .fh-brand{{font-size:16px;font-weight:700;letter-spacing:.06em;color:#1a1a1a}}.folha-head .fh-brand b{{font-weight:700;color:#9c7b2e}}
.folha-head .fh-sub{{font-size:7px;color:#9c7b2e;text-transform:uppercase;letter-spacing:.05em;margin-top:2px}}
.fh-right{{text-align:right}}.fh-right .fh-tt{{font-size:11px;color:#9c7b2e;font-weight:700}}.fh-right .fh-dt{{font-size:7px;color:#999;margin-top:1px}}
.fh-media{{font-size:9px;color:#1a1a1a;font-weight:700;margin-top:2px}}
table.fesp{{width:100%;border-collapse:separate;border-spacing:3px;margin-top:8px}}
table.fesp th{{font-size:7.5px;color:#857c68;text-transform:uppercase;letter-spacing:.03em;padding:2px;text-align:center;font-weight:500}}
table.fesp th small{{display:block;font-size:6.5px;color:#a59c88;font-weight:400}}
table.fesp .fand{{color:#7d6021;font-weight:700;font-size:9px;text-align:center}}
.fcell{{border-radius:5px;padding:4px 3px;text-align:center}}
.fcell .fa{{display:block;font-weight:700;font-size:10px;color:#2c2719}}
.fcell .fs{{display:block;font-size:6.5px;margin-top:0}}
.fcell .fm{{display:block;font-size:7px;font-weight:600;margin-top:1px;color:#5a5443}}
.f-ok{{background:rgba(31,122,77,0.15);border:1px solid rgba(31,122,77,0.45)}}.f-ok .fs{{color:#1c6a40}}
.f-res{{background:rgba(194,118,47,0.17);border:1px solid rgba(194,118,47,0.50)}}.f-res .fs{{color:#955310}}
.f-sold{{background:rgba(155,82,82,0.15);border:1px solid rgba(155,82,82,0.45)}}.f-sold .fs{{color:#8a3b3b}}
.folha-legend{{display:flex;gap:14px;margin-top:8px;font-size:8px;color:#6a6253}}
.folha-legend i{{display:inline-block;width:9px;height:9px;border-radius:2px;margin-right:4px;vertical-align:middle}}
.folha-foot{{margin-top:9px;border-top:.5px solid #cfc6b3;padding-top:5px;font-size:6.5px;color:#857c68;text-align:center}}
/* folha da TABELA (paisagem) */
#folha-tabela{{page-break-inside:avoid}}
.ftbl-blk{{background:#f4ecd8;color:#7d6021;border-left:3px solid #b9974a;font-size:7.3px;font-weight:700;letter-spacing:.03em;padding:3px 7px;margin:6px 0 0;text-transform:uppercase}}
table.ftbl{{width:100%;border-collapse:collapse;font-size:6.3px;margin-bottom:1px}}
table.ftbl th{{background:#3D4E45;color:#fff;font-weight:500;font-size:5.7px;text-transform:uppercase;letter-spacing:.02em;padding:2px 4px;text-align:right;border-bottom:1px solid #b9974a}}
table.ftbl th.l{{text-align:left}}
table.ftbl td{{padding:1.5px 4px;border-bottom:.4px solid #d8d0bf;color:#2c2719;text-align:right;font-variant-numeric:tabular-nums}}
table.ftbl td.l{{text-align:left}}table.ftbl td.b{{font-weight:700}}
table.ftbl tr:nth-child(even) td{{background:#f7f3ea}}
table.ftbl tr{{page-break-inside:avoid}}
.ftbl .pill{{font-size:6px;padding:1px 5px;border-radius:20px;font-weight:600;border:1px solid}}
.ftbl .p-ok{{color:#1c6a40;background:rgba(31,122,77,0.12);border-color:rgba(31,122,77,0.4)}}
.ftbl .p-res{{color:#955310;background:rgba(194,118,47,0.14);border-color:rgba(194,118,47,0.45)}}
.ftbl-obs{{margin-top:8px;font-size:6.5px;color:#857c68;border-top:.5px solid #cfc6b3;padding-top:5px}}
.ftbl-obs b{{color:#7d6021;display:block;margin-bottom:3px}}
@media print{{
  *{{-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  #gate,nav.tabs,.printbtn,.btnrow{{display:none!important}}
  body{{background:#fff!important;color:#2c2719!important}}
  body.exporting #app{{display:none!important}}
  body.exporting #folhas{{display:block!important}}
  #folhas .folha{{display:none}} #folhas .folha.active{{display:block}}
  #folhas .folha,#folhas table.fesp,#folhas table.fesp tr{{page-break-inside:avoid}}
  body:not(.exporting) header.top,body:not(.exporting) footer,body:not(.exporting) .legend{{display:none!important}}
  body:not(.exporting) #app{{display:block!important;max-width:none;margin:0;padding:0}}
  body:not(.exporting) section.view{{display:none!important}}
  body:not(.exporting) section.view.printing{{display:block!important;padding:0}}
  body:not(.exporting) h2.vt,body:not(.exporting) p.vd{{display:none!important}}
  .printhdr{{display:flex!important}} .printftr{{display:block!important}}
  .blocohdr{{background:#16140f!important;color:#c2a15a!important;border-left:3px solid #b9974a;border-radius:0;font-size:9pt;padding:5px 10px;margin:10px 0 0}}
  table.ptable{{font-size:7.3pt;width:100%}}
  table.ptable th{{background:#2a261d!important;color:#fff!important;border-bottom:1px solid #b9974a;padding:4px 5px;font-size:6.5pt}}
  table.ptable td{{border-bottom:.4px solid #cfc6b3;color:#2c2719!important;background:#fff!important;padding:3px 5px}}
  table.ptable tr:nth-child(even) td{{background:#f6f2ea!important}}
  tr.st-ok .pill{{color:#1c6a40!important;background:#fff!important;border-color:#1c6a40!important}}
  tr.st-res .pill,tr.st-sold .pill{{color:#9a5310!important;background:#fff!important;border-color:#9a5310!important}}
  .obs{{background:#f6f2ea!important;color:#6a6253!important;border:1px solid #cfc6b3}}.obs b{{color:#7d6021!important}}
}}
@media(max-width:680px){{table.ptable{{font-size:10.5px}}.ecell .erm2{{font-size:10px}}}}
</style>
<style id="pagestyle">@page{{size:A4 landscape;margin:11mm 9mm}}</style></head>
<body>

<div id="gate">
  <div class="brand">DOM MANUEL</div>
  <div class="sub">APRESENTAÇÃO · ACESSO RESTRITO</div>
  <form onsubmit="return tryEnter(event)">
    <input id="pw" type="password" placeholder="Senha" autocomplete="off" autofocus>
    <button type="submit">Entrar</button>
  </form>
  <div class="err" id="err"></div>
</div>

<div id="app">
  <header class="top">
    <div class="brand">DOM <b>MANUEL</b></div>
    <div class="meta">{esc(CONFIG['subtitulo'])}<br>Tabela vigente: {esc(vigente)}<br>Atualizado em {esc(atual or '—')}</div>
  </header>

  <nav class="tabs">
    <button class="on" data-view="book">Book de Vendas</button>
    <button data-view="pesquisa">Pesquisa de Mercado</button>
    <button data-view="espelho">Espelho</button>
    <button data-view="tabela">Tabela {esc(vigente)}</button>
    <button data-view="institucional">Institucional DOM</button>
  </nav>

  <section class="view on" data-name="book">
    <h2 class="vt">Book de Vendas</h2>
    <p class="vd">Apresentação comercial do empreendimento.</p>
    <div class="btnrow">
      <button class="act" onclick="present('bookframe')">▶ Apresentar em tela cheia</button>
      <a class="act ghost" href="book.html" target="_blank" rel="noopener">Abrir em nova aba</a>
    </div>
    <iframe id="bookframe" class="bookframe" src="book.html" title="Book Dom Manuel" allowfullscreen></iframe>
  </section>

  <section class="view" data-name="institucional">
    <h2 class="vt">Institucional DOM</h2>
    <p class="vd">Apresentação institucional da DOM Incorporação.</p>
    <div class="btnrow">
      <button class="act" onclick="present('instframe')">▶ Apresentar em tela cheia</button>
      <a class="act ghost" href="institucional_dom.pdf" target="_blank" rel="noopener">Abrir em nova aba</a>
    </div>
    <iframe id="instframe" class="bookframe" src="institucional_dom.pdf" title="Institucional DOM" allowfullscreen></iframe>
  </section>

  <section class="view" data-name="espelho">
    <button class="printbtn" onclick="printFolha('espelho')">Exportar PDF</button>
    <h2 class="vt">Espelho de Vendas</h2>
    <p class="vd">Disponibilidade das 45 unidades · tabela vigente ({esc(vigente)}) · preço médio {esc(media)}.</p>
    <button class="toggleprice" id="btn-rm2" onclick="togglePr('rm2')">R$/m²</button>
    <button class="toggleprice" id="btn-ticket" onclick="togglePr('ticket')">Ticket</button>
    <div class="legend">{legenda}</div>
    {esp}
  </section>

  <section class="view" data-name="tabela">
    <button class="printbtn" onclick="printFolha('tabela')">Exportar PDF</button>
    <div class="printhdr"><div><div class="ph-brand">DOM <b>MANUEL</b></div><div class="ph-sub">Condomínio Dom Manuel · Ponta D'Areia — São Luís/MA</div></div>
      <div class="ph-right"><span>Tabela Comercial · {esc(vigente)}</span><small>Régua {esc(CONFIG['regua_versao'])} · atualizado {esc(atual or '—')}</small></div></div>
    <h2 class="vt">Tabela Comercial · {esc(vigente)}</h2>
    <p class="vd">Plano de pagamento por unidade · valores em reais · preço médio {esc(media)}.</p>
    {tabela}
    <div class="printftr">Tabela para fins de reserva, sujeita a alteração sem aviso · parcelas corrigidas pelo INCC na obra · 60% financiado no habite-se · valores sujeitos a análise de crédito · DOM Incorporação</div>
  </section>

  <section class="view" data-name="pesquisa">
    <h2 class="vt">Pesquisa de Mercado</h2>
    <p class="vd">Estudo de inteligência de mercado · Condomínio Dom Manuel.</p>
    <div class="btnrow">
      <button class="act" onclick="present('pesqframe')">▶ Apresentar em tela cheia</button>
      <a class="act ghost" href="pesquisa_mercado.html" target="_blank" rel="noopener">Abrir em nova aba</a>
    </div>
    <iframe id="pesqframe" class="bookframe" src="pesquisa_mercado.html" title="Pesquisa de Mercado" allowfullscreen></iframe>
  </section>

  <footer>Portal gerado automaticamente do painel + régua · fonte única DOM</footer>
</div>

<div id="folhas"><div class="folha" id="folha-espelho">{folha_esp}</div><div class="folha" id="folha-tabela">{folha_tab}</div></div>

<script>
var SENHA = {CONFIG['senha']!r};
function tryEnter(e){{e.preventDefault();
  if(document.getElementById('pw').value===SENHA){{
    document.getElementById('gate').style.display='none';
    document.getElementById('app').style.display='block';
  }} else {{document.getElementById('err').textContent='Senha incorreta.';}}
  return false;}}
document.querySelectorAll('nav.tabs button[data-view]').forEach(function(b){{
  b.onclick=function(){{
    document.querySelectorAll('nav.tabs button').forEach(x=>x.classList.remove('on'));
    b.classList.add('on');
    var n=b.getAttribute('data-view');
    document.querySelectorAll('section.view').forEach(function(s){{
      s.classList.toggle('on', s.getAttribute('data-name')===n);}});
  }};
}});
function present(id){{
  var f=document.getElementById(id);
  if(f.requestFullscreen) f.requestFullscreen();
  else if(f.webkitRequestFullscreen) f.webkitRequestFullscreen();
  else window.open(f.src,'_blank');
}}
function togglePr(which){{
  var cls=which==='rm2'?'show-rm2':'show-ticket';
  var on=document.body.classList.toggle(cls);
  document.getElementById('btn-'+which).classList.toggle('on',on);
}}
function printFolha(which){{
  var ps=document.getElementById('pagestyle');
  var marg=(which==='tabela')?'11mm 8mm':'16mm 14mm';
  ps.textContent='@page{{size:A4 portrait;margin:'+marg+'}}';
  var nomes={{espelho:'Espelho Dom Manuel',tabela:'Tabela Dom Manuel'}};
  var t0=document.title; if(nomes[which]) document.title=nomes[which];
  document.body.classList.add('exporting');
  document.querySelectorAll('#folhas .folha').forEach(function(f){{
    f.classList.toggle('active', f.id==='folha-'+which);}});
  window.print();
  setTimeout(function(){{document.body.classList.remove('exporting');
    ps.textContent='@page{{size:A4 landscape;margin:11mm 9mm}}';document.title=t0;}},500);
}}
</script>
</body></html>"""

# ───────────────────────── MAIN ─────────────────────────
def main():
    status_by_apto, atual = ler_painel()
    vigente = CONFIG["vigente"]
    C = TEMAS.get(CONFIG["tema"], TEMAS["contraste"])
    tab = ler_tabela(vigente)
    out = render(status_by_apto, atual, tab, vigente, C)
    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write(out)
    # resumo
    from collections import Counter
    cont = Counter(mascara(s) for s in status_by_apto.values())
    print(f"OK -> {OUT}")
    print(f"  tema: {CONFIG['tema']} · mostrar_vendido: {CONFIG['mostrar_vendido']}")
    print(f"  tabela vigente: {vigente} · {len(tab['blocks'])} blocos · {len(tab['idx'])} unidades")
    print(f"  espelho: {dict(cont)}")

if __name__ == "__main__":
    main()
